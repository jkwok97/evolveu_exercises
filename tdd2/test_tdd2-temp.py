import unittest
import tdd2
import datetime
from openpyxl import Workbook, worksheet, load_workbook

class TestDates(unittest.TestCase):

	def test_my_max(self):
		self.assertEqual(5,tdd2.my_max([1,2,3,4,5]))
		self.assertEqual(25,tdd2.my_max([1,25,3,4,5]))
		self.assertEqual(32,tdd2.my_max([1,32,3,4,5]))
		self.assertEqual(11,tdd2.my_max([11,2,3,4,5]))

	def test_my_min(self):
		self.assertEqual(1,tdd2.my_min([1,2,3,4,5]))
		self.assertEqual(2,tdd2.my_min([11,2,3,4,5]))
		self.assertEqual(3,tdd2.my_min([19,25,3,4,5]))
		self.assertEqual(4,tdd2.my_min([13,27,31,4,5]))

	def test_has_string(self):
		self.assertEqual(["Mary had"],tdd2.has_string(["Mary had","a little lamb","little lamb","Whose fleece",],"Mary"))

class TestTddExcel(unittest.TestCase):	

	def test_to_date(self):

		dt = tdd2.to_date("2018-09-03")
		self.assertIsInstance(dt, datetime.date)
		self.assertEqual(2018,dt.year)
		self.assertEqual(9,dt.month)
		self.assertEqual(3,dt.day)

	def test_days_between(self):
		self.assertEqual(1, tdd2.date_diff("2018-09-02", "2018-09-01"))
		self.assertEqual(2080, tdd2.date_diff("2018-09-01", "2012-12-21"))

	def test_value_contained(self):
		self.assertTrue(tdd2.contains(['a','b','d'],"a"))

	def test_added_content(self):
		self.assertEqual(6, tdd2.add_contents([1,2,3]))

	def test_add_to_string(self):
		self.assertEqual('one mine', tdd2.lookupx({1:'one', 2:'two', 3:'three'} , 1))

## Major Assignment

	def test_find_tab(self):
		wb = load_workbook('cSpace_Bookingv1.xlsx')
		entered_date = '2018-07'

		ws = tdd2.find_tab(wb, entered_date)
		self.assertIsInstance(ws, worksheet.Worksheet)
		self.assertEqual(entered_date[0:7], ws.title)

		self.assertIsNone(tdd2.find_tab(wb, '2011-07'))

	def test_cell_has_value(self):
		wb = Workbook()
		ws = wb.active

		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = ''
		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = 'Room a'
		self.assertTrue(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = '   '
		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		self.assertFalse(tdd2.cell_has_value(ws, 100, 100))
		self.assertFalse(tdd2.cell_has_value(ws, 1, 100))
		self.assertFalse(tdd2.cell_has_value(ws, 100, 1))

	def test_verify_name(self):
		self.assertTrue(True, tdd2.lookup({'client_1':'Carri', 'client_2':'Sherrie', 'client_3':'Geri'} , 'Carri'))




