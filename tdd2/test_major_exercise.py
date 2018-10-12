
import unittest
import major_exercise
import datetime
from openpyxl import Workbook, worksheet, load_workbook


class TestTddExcel(unittest.TestCase):

	def test_find_tab(self):
		wb = load_workbook('cSpace_Bookingv1.xlsx')
		entered_date = '2018-07'

		ws = major_exercise.find_tab(wb, entered_date)
		self.assertIsInstance(ws, worksheet.Worksheet)
		self.assertEqual(entered_date[0:7], ws.title)

		self.assertIsNone(major_exercise.find_tab(wb, '2011-07'))

	def test_cell_has_value(self):
		wb = Workbook()
		ws = wb.active

		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = ''
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = 'Room a'
		self.assertTrue(major_exercise.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = '   '
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 1))

		self.assertFalse(major_exercise.cell_has_value(ws, 100, 100))
		self.assertFalse(major_exercise.cell_has_value(ws, 1, 100))
		self.assertFalse(major_exercise.cell_has_value(ws, 100, 1))

	def test_verify_name(self):
		self.assertTrue(True, major_exercise.lookup({'client_1':'Carri', 'client_2':'Sherrie', 'client_3':'Geri'} , 'Carri'))

class Testoo1_exercise(unittest.TestCase):

	def test_client_list(self):
		client_1 = major_exercise.Clients('Carri', 'Cordon')
		client_2 = major_exercise.Clients('Caren', 'Alsop')
		client_3 = major_exercise.Clients('Sierra', 'Sandoval')
		client_4 = major_exercise.Clients('Delcie', 'Montalvan')
		client_5 = major_exercise.Clients('Geri', 'Gallant')
		client_6 = major_exercise.Clients('Lavera', 'Platter')
		client_7 = major_exercise.Clients('Soila', 'Cozad')
		client_8 = major_exercise.Clients('Valeri', 'Miranda')
		client_9 = major_exercise.Clients('Samira', 'Harlin')
		client_10 = major_exercise.Clients('Leone', 'Elwell')
		client_11 = major_exercise.Clients('Marylee', 'Shoener')
		client_12 = major_exercise.Clients('Annalisa', 'Jansen')
		client_13 = major_exercise.Clients('Dorathy', 'Koprowski')
		client_14 = major_exercise.Clients('Denese', 'Askew')
		client_15 = major_exercise.Clients('Larraine', 'Primrose')
		client_16 = major_exercise.Clients('Cassie', 'Close')
		client_17 = major_exercise.Clients('Sherrie', 'Iraheta')
		client_18 = major_exercise.Clients('Alicia', 'Baskin')
		client_19 = major_exercise.Clients('Jeramy', 'Beeler')
		client_20 = major_exercise.Clients('Clarissa', 'Petterson')

		self.assertEqual('Carri Cordon', client_1.name)
		self.assertEqual('Leone Elwell', client_10.name)
		self.assertEqual('Sherrie Iraheta', client_17.name)
		self.assertEqual('Geri Gallant', client_5.name)