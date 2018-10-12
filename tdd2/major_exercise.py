
import sys
import openpyxl
from datetime import datetime

# Exercise 1

wb = openpyxl.load_workbook('cSpace_Bookingv1.xlsx')
sheets = wb['Clients']

names = []
issues = []
for c in sheets.rows:
	names.append(c[0].value)#
	issues.append(c[5].value)
del names[0]
for count, name in enumerate (names,1):
	print (name.split()[0], name.split()[1], issues[count])

# exercise 2

def find_tab(wb, date_s):
	for sheet in wb:
		if sheet.title == date_s[0:7]:
			return sheet
		return None

def cell_has_value(ws, row, col):
	if row > ws.max_row:
		return False
	if col > ws.max_column:
		return False
	if not ws.cell(row=row, column=col).value:
		return False
	if not ws.cell(row=row, column=col).value.strip():
		return False
	return True

def lookup(dictionary, key_value):
	if key_value in dictionary.values():
		return True
	return False

# exercise 3

class Clients:

# create an instance of the class Client from the excel worksheet Client
# take that instance and put inside a dictionary
# get all data from a month tab
# get all data from facilities tab
# get all data from rates tab
# set rate for each type of space
# add rate for each client 

#	client_list = {}

	def __init__ (self, first, last):
		self.first = first
		self.last = last

	def name(self):
		return '{} {}'.format(self.first, self.last)

#client_1 = Clients.name
#print (client_1)



#	def __iters__(self):
#		iters = dict((x,y) for x,y in Clients.__dict__.items())
#		iters.update(self.__dict)
#		for x,y in iters.items():
#			yield x,y

#client_1 = Clients()
#print(dict(client_1))