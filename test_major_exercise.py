
import unittest
import major_exercise
import datetime
from openpyxl import Workbook, worksheet, load_workbook


class TestTddExcel(unittest.TestCase):

	def test_find_tab(self):
		wb = load_workbook('cSpace_Bookingv1.xlsx')
		entered_date = '2018-07'

		ws = tdd2.find_tab(wb, entered_date)
		self.assertIsInstance(ws, worksheet.Worksheet)
		self.assertEqual(entered_date[0:7], ws.title)

		self.assertIsNone(tdd2.find_tab(wb, '2011-07'))

	def test_cell_has_value(self):
		wb = Workbook()
		ws = wb.active

		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = ''
		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = 'Room a'
		self.assertTrue(tdd2.cell_has_value(ws, 1, 1))

		ws.cell(row=1, column=1).value = '   '
		self.assertFalse(tdd2.cell_has_value(ws, 1, 1))

		self.assertFalse(tdd2.cell_has_value(ws, 100, 100))
		self.assertFalse(tdd2.cell_has_value(ws, 1, 100))
		self.assertFalse(tdd2.cell_has_value(ws, 100, 1))

	def test_verify_name(self):
		self.assertTrue(True, tdd2.lookup({'client_1':'Carri', 'client_2':'Sherrie', 'client_3':'Geri'} , 'Carri'))

class Testoo1_exercise(unittest.TestCase):

	def test_client_list(self):
		client_1 = oo1.Clients('Carri', 'Cordon')
		client_2 = oo1.Clients('Caren', 'Alsop')
		client_3 = oo1.Clients('Sierra', 'Sandoval')
		client_4 = oo1.Clients('Delcie', 'Montalvan')
		client_5 = oo1.Clients('Geri', 'Gallant')
		client_6 = oo1.Clients('Lavera', 'Platter')
		client_7 = oo1.Clients('Soila', 'Cozad')
		client_8 = oo1.Clients('Valeri', 'Miranda')
		client_9 = oo1.Clients('Samira', 'Harlin')
		client_10 = oo1.Clients('Leone', 'Elwell')
		client_11 = oo1.Clients('Marylee', 'Shoener')
		client_12 = oo1.Clients('Annalisa', 'Jansen')
		client_13 = oo1.Clients('Dorathy', 'Koprowski')
		client_14 = oo1.Clients('Denese', 'Askew')
		client_15 = oo1.Clients('Larraine', 'Primrose')
		client_16 = oo1.Clients('Cassie', 'Close')
		client_17 = oo1.Clients('Sherrie', 'Iraheta')
		client_18 = oo1.Clients('Alicia', 'Baskin')
		client_19 = oo1.Clients('Jeramy', 'Beeler')
		client_20 = oo1.Clients('Clarissa', 'Petterson')

		self.assertEqual('Carri Cordon', client_1.name)
		self.assertEqual('Leone Elwell', client_10.name)
		self.assertEqual('Sherrie Iraheta', client_17.name)
		self.assertEqual('Geri Gallant', client_5.name)