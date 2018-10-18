import openpyxl
import sys

from dateutil import parser


def empty_room(booking_date):

    return ['Desk 3rd floor 3', 'Desk 3rd floor 5', 'Desk 3rd floor 6', 'Desk 3rd floor 7', 'Desk 3rd floor 9']

def get_year_month(bookingdate):
    
    return bookingdate[0:7]

def get_ymd(bookingdate):
	if bookingdate[8:]=='01':
		ymd = parser.parse(bookingdate)
		#ymd = bookingdate[0:4] + '/' + bookingdate[5:7] + '/' + bookingdate[8:]
	elif (int(bookingdate[8:]) <= 10) and (int(bookingdate[8:]) >= 2):
		ymd = '=A' + bookingdate[9] + '+1'
	else:
		ymd = '=A' + bookingdate[8:] + '+1'
	return ymd


def getSheetName(filename):
    workbook = openpyxl.load_workbook(filename)
    return workbook.sheetnames

def getrooms(bookingdate):
	roomlist = []
	filename = 'cSpace_Bookingv1.xlsx'

	workbook = openpyxl.load_workbook(filename)
	allsheets = getSheetName(filename)
	ws = workbook[get_year_month(bookingdate)]
	
	dateList = []
	for cell in ws['A']:
		dateList.append(cell.value)

	selectRow = dateList.index(get_ymd(bookingdate)) + 1
	a = dateList[selectRow]

	rowList = [] 
	
	roomList = [] # list value in every cell
	
	fullrooms = []

	for cell in ws[selectRow]:
		rowList.append(cell.value)

	for cell in ws[1]:
		roomList.append(cell.value)


	for i in range(len(rowList)):
		if rowList[i]:
			fullrooms.append(roomList[i])
	
	for room in fullrooms:
		roomList.remove(room)

		# for i in range(len(rowList)):
	# 	if rowList[i]:
	# 		del roomList[i]

	return roomList

date=input("Please input the date using the format yyyy-mm-dd")
print(getrooms(date))



