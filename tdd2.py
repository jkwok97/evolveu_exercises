import sys
from openpyxl import load_workbook
from datetime import datetime

def add_five(num):
	return num + 5

def my_max(largest_number):
	if largest_number != []:
		largest = largest_number[0]
		for large in largest_number:
			if large > largest:
				largest=large
		return largest
	else:
		return None
#	largest_number.sort()
#	return (largest_number[-1])

def my_min(smallest_number):
	if smallest_number != []:
		smallest = smallest_number[0]
		for small in smallest_number:
			if small < smallest:
				smallest=small
		return smallest
	else:
		return None
#	smallest_number.sort()
#	return (smallest_number[0])

def has_string(show_list, show_string):
	list_string = []
	for item in show_list:
		if show_string in item:
			list_string.append(item)
	return (list_string)

def to_date(dt):
	return datetime.strptime(dt, "%Y-%m-%d").date()

def date_diff(date_1, date_2):
	date1 = to_date(date_1)
	date2 = to_date(date_2)
	return (abs(date1 - date2).days)

# def end_date(today, mayan):
# 	date3 = to_date(today)
# 	date4 = to_date(mayan)
# 	return (abs(date3 - date4).days)

def contains(list1, string):
	if string in list1:
		return True
	else:
		return False

def add_contents(list_of_nums):
    sum_numbers = 0
    for x in list_of_nums:
        sum_numbers += x
    return sum_numbers
#	return (sum(list_of_nums))

def lookupx(dictionary, key_place):
	if key_place in dictionary:
		key_value = dictionary.get(key_place)
		key_value = (key_value + ' mine')
		return key_value
	else:
		return (' mine')

## Major Assignment

def find_tab(wb, date_s):
	for sheet in wb:
		if sheet.title == date_s[0:7]:
			return sheet
		return None

def cell_has_value(ws, row, col):
	if row > ws.max_row:
		return False
	if col > ws.max_column:
		return False
	if not ws.cell(row=row, column=col).value:
		return False
	if not ws.cell(row=row, column=col).value.strip():
		return False
	return True

def lookup(dictionary, key_value):
	if key_value in dictionary.values():
		return True
	return False

